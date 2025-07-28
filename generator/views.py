from django.shortcuts import render
from django.http import JsonResponse
from .openAI_api import generate_test_cases
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font
from django.http import HttpResponse


def home(request):
    """Display the input form for entering requirements"""
    return render(request, "generator/input_form.html")


def generate_testcases(request):
    """Generate test cases based on user requirements"""
    if request.method == "POST":
        requirement = request.POST.get("requirement")

        if not requirement:
            return render(
                request,
                "generator/input_form.html",
                {"error": "Please enter a requirement."},
            )

        # Generate test cases using the API
        testcases = generate_test_cases(requirement)

        # Store in session for JSON endpoints
        request.session["testcases"] = testcases
        request.session["requirement"] = requirement

        return render(
            request,
            "generator/result.html",
            {"requirement": requirement, "testcases": testcases},
        )

    # If GET request, redirect to home
    return render(request, "generator/input_form.html")


def result(request):
    """Display stored test cases from session"""
    testcases = request.session.get("testcases", [])
    requirement = request.session.get("requirement", "")

    if not testcases:
        return render(
            request,
            "generator/input_form.html",
            {"error": "No test cases found. Please generate test cases first."},
        )

    return render(
        request,
        "generator/result.html",
        {"testcases": testcases, "requirement": requirement},
    )


def test_cases_json(request):
    """Return all test cases as JSON"""
    if "testcases" in request.session:
        return JsonResponse(request.session["testcases"], safe=False)
    else:
        return JsonResponse({"error": "No test cases found in session"}, status=404)


def test_case_json(request, case_id):
    """Return a specific test case as JSON"""
    if "testcases" in request.session:
        testcases = request.session["testcases"]
        for case in testcases:
            if case.get("id") == int(case_id):
                return JsonResponse(case)
        return JsonResponse(
            {"error": f"Test case with ID {case_id} not found"}, status=404
        )
    else:
        return JsonResponse({"error": "No test cases found in session"}, status=404)


def export_testcases_excel(request):
    testcases = request.session.get("testcases", [])
    if not testcases:
        return HttpResponse("No test cases to export.", status=400)

    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = "Test Cases"

    # Main table headers
    headers = [
        "ID",
        "Title",
        "Description",
        "Type",
        "Priority",
        "Expected Output",
        "Manual Steps",
        "Pytest Code",
        "Robot Code",
    ]
    ws_main.append(headers)

    # Create code sheet
    ws_code = wb.create_sheet(title="Code Snippets")
    ws_code.append(["ID", "Type", "Code Type", "Code"])

    # Fill main table and code sheet
    for idx, case in enumerate(testcases, start=2):
        pytest_cell = f'=HYPERLINK("#\'Code Snippets\'!D{idx*2-1}", "View Pytest Code")'
        robot_cell = f'=HYPERLINK("#\'Code Snippets\'!D{idx*2}", "View Robot Code")'
        ws_main.append(
            [
                case.get("id", ""),
                case.get("title", ""),
                case.get("description", ""),
                case.get("type", ""),
                case.get("priority", ""),
                case.get("expected_output", ""),
                case.get("manual_steps", ""),
                pytest_cell,
                robot_cell,
            ]
        )
        # Add code to code sheet (one row for each code type)
        ws_code.append(
            [
                case.get("id", ""),
                case.get("title", ""),
                "Pytest",
                case.get("pytest_code", ""),
            ]
        )
        ws_code.append(
            [
                case.get("id", ""),
                case.get("title", ""),
                "Robot Framework",
                case.get("robot_code", ""),
            ]
        )

    # Make hyperlinks blue and underlined
    for row in ws_main.iter_rows(min_row=2, min_col=8, max_col=9):
        for cell in row:
            cell.font = Font(color="0000FF", underline="single")

    # Set column widths for readability
    for ws in [ws_main, ws_code]:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="test_cases.xlsx"'
    wb.save(response)
    return response
